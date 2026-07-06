import os
import pandas as pd

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

from openpyxl.formatting.rule import ColorScaleRule


OUTPUT = "output/screener_output.xlsx"


def export_screeners(results):

    os.makedirs("output", exist_ok=True)

    with pd.ExcelWriter(
        OUTPUT,
        engine="openpyxl"
    ) as writer:

        for sheet_name, df in results.items():

            cols = [

                "overall_rank",
                "sector_rank",
                "company_id",
                "broad_sector",
                "year",
                "rating",
                "percentile",
                "composite_quality_score",
                "return_on_equity_pct",
                "net_profit_margin_pct",
                "debt_to_equity",
                "interest_coverage",
                "asset_turnover",
                "free_cash_flow_cr"

            ]

            cols = [c for c in cols if c in df.columns]

            df = df.sort_values(
                "overall_rank"
            )

            df[cols].to_excel(

                writer,

                sheet_name=sheet_name[:31],

                index=False

            )

            ws = writer.sheets[sheet_name[:31]]

            # ---------------------------------
            # Freeze Header
            # ---------------------------------

            ws.freeze_panes = "A2"

            # ---------------------------------
            # Auto Filter
            # ---------------------------------

            ws.auto_filter.ref = ws.dimensions

            # ---------------------------------
            # Header Formatting
            # ---------------------------------

            header_fill = PatternFill(

                fill_type="solid",

                start_color="1F4E78",

                end_color="1F4E78"

            )

            header_font = Font(

                bold=True,

                color="FFFFFF"

            )

            thin = Side(

                style="thin",

                color="DDDDDD"

            )

            for cell in ws[1]:

                cell.fill = header_fill

                cell.font = header_font

                cell.alignment = Alignment(

                    horizontal="center",

                    vertical="center"

                )

                cell.border = Border(

                    left=thin,

                    right=thin,

                    top=thin,

                    bottom=thin

                )

            # ---------------------------------
            # Auto Width
            # ---------------------------------

            for column_cells in ws.columns:

                length = max(

                    len(str(cell.value))

                    if cell.value is not None

                    else 0

                    for cell in column_cells

                )

                ws.column_dimensions[

                    column_cells[0].column_letter

                ].width = length + 4

            # ---------------------------------
            # Color Scale for Composite Score
            # ---------------------------------

            if (
                "composite_quality_score" in cols
                and len(df) >= 2
            ):

                col = cols.index("composite_quality_score") + 1

                letter = ws.cell(
                    row=1,
                    column=col
                ).column_letter

                rng = f"{letter}2:{letter}{ws.max_row}"

                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8696B",

                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",

                        end_type="max",
                        end_color="63BE7B"
                    )
                )

    print(f"\nSaved to {OUTPUT}")